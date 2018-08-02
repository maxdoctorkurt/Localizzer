#!/usr/bin/env python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook
import codecs
import os
import re

def columnsToRows(colsList):

	rows = len(colsList[0])
	cols = len(colsList);

	result = [["null" for c in range(0, cols)] for r in range(0, rows)]

	for i in range(0, rows):
		for j in range(0, cols):
			result[i][j] = (colsList[j][i]).value

	return result

def fileFromStr(fname, string):
	f = codecs.open(fname, 'w', "utf-8")
	f.write(string)
	f.close()

def makeDir(dirname):
	if not os.path.exists(dirname):
		os.makedirs(dirname)

def escapeStringAndroid(string):

	return str(string).translate(str.maketrans({ 
                                       	"\"": r"&quot;",
                                        "\'": r"&apos;",
                                        "&": r"&amp;",
                                        "<": r"&lt;",
                                        ">": r"&gt;"
                                        }))

def genLocFiles(rows, fnamePrefix):

	hat = "<resources>\n\r"
	footer = "</resources>"

	langs = rows[0][1:]

	for lang in langs:

		stringsAndroid = ""
		stringsIos = ""

		for row in rows[1:]:

			key = row[0]
			valAndroid = escapeStringAndroid(row[rows[0].index(lang)])
			valIos = str(row[rows[0].index(lang)])

			stringsAndroid = stringsAndroid + "".join(["\t<string name=\"", key, "\">", valAndroid, "</string>\n\r"])
			stringsIos = stringsIos + "".join(["\t\"", key, "\" = \"", valIos, "\";\n\r"])

		androidDir = "android"
		iosDir = "ios"

		makeDir(androidDir)
		fileFromStr(os.path.join(androidDir, "".join([fnamePrefix, "_", lang, ".xml"])), "".join([hat, stringsAndroid, footer]))

		makeDir(iosDir)
		fileFromStr(os.path.join(iosDir, "".join([fnamePrefix, "_", lang, ".strings"])), stringsIos)

def genLocalizations():

	wb = load_workbook(filename = 'loc.xlsx')

	for sname in wb.sheetnames:
		sheet = wb[sname]

		maxr = sheet.max_row
		maxc = sheet.max_column

		columns = []

		az = ["a", "b", "c", "d", "e"]

		for columnName in az:

			if(len(columns) >= maxc):
				break
			column = sheet[columnName]
			columns.append(column)

		rows = columnsToRows(columns)

		genLocFiles(rows, sname)


genLocalizations()